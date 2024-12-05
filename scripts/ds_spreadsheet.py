from time import sleep

from django.contrib.sites.models import Site
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

from recoco.apps.demarches_simplifiees.models import DSMapping, DSResource, Field
from recoco.apps.demarches_simplifiees.tasks import load_ds_resource_schema
from recoco.apps.geomatics.models import Department
from recoco.apps.resources.models import Resource


def main():
    site = Site.objects.filter(domain="sosponts.recoconseil.fr").first()

    resource = Resource.objects.get(pk=487)

    ds_resource, created = DSResource.objects.get_or_create(
        name="pnp-travaux-dispositif-d-aide",
        defaults={"resource": resource},
    )
    if created:
        ds_resource.departments.add(Department.objects.all())
        load_ds_resource_schema.delay(ds_resource.id)
        sleep(5)
        ds_resource.refresh_from_db()

    ds_mapping, _ = DSMapping.objects.get_or_create(
        ds_resource=ds_resource,
        site=site,
        defaults={"mapping": {}},  # TODO: prefill the mapping with existing values
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "mapping"

    # config
    max_rows = 1000
    default_col_width = 60
    ds_fields_sheet_name = "ds-fields"
    lookup_fields_sheet_name = "lookup-fields"

    # create the mapping sheet headers
    header_fill = PatternFill(start_color="FFC7CE", fill_type="solid")
    ws["A1"] = "Nom du champ DS"
    ws["A1"].fill = header_fill
    ws.column_dimensions["A"].width = default_col_width
    ws["B1"] = "Nom du champ Recoco"
    ws["B1"].fill = header_fill
    ws.column_dimensions["B"].width = default_col_width

    ds_fields: list[Field] = ds_mapping.ds_fields
    lookup_fields: list[Field] = ds_mapping.lookup_fields

    # create a sheet to register the ds fields dropdown values
    ws_ds_fields = wb.create_sheet(ds_fields_sheet_name)
    ws_ds_fields.column_dimensions["A"].width = default_col_width
    ws_ds_fields.column_dimensions["B"].width = default_col_width
    for i, f in enumerate(ds_fields):
        ws_ds_fields[f"A{i+1}"] = f.id
        ws_ds_fields[f"B{i+1}"] = f.label

    # create a sheet to register the lookup fields dropdown values
    ws_lookup_fields = wb.create_sheet(lookup_fields_sheet_name)
    ws_lookup_fields.column_dimensions["A"].width = default_col_width
    ws_lookup_fields.column_dimensions["B"].width = default_col_width
    for i, f in enumerate(lookup_fields):
        ws_lookup_fields[f"A{i+1}"] = f.id
        ws_lookup_fields[f"B{i+1}"] = f.label

    # setup the dropdowns for the mapping sheet
    ds_field_validation = DataValidation(
        type="list",
        formula1=f"'{ds_fields_sheet_name}'!$B$1:$B${len(ds_fields)}",
        allowBlank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    ws.add_data_validation(ds_field_validation)
    for row in range(2, max_rows):
        ds_field_validation.add(ws[f"A{row}"])

    recoco_field_validation = DataValidation(
        type="list",
        formula1=f"'{lookup_fields_sheet_name}'!$A$1:$A${len(lookup_fields)}",
        allow_blank=True,
        showInputMessage=True,
        showErrorMessage=True,
    )
    ws.add_data_validation(recoco_field_validation)
    for row in range(2, max_rows):
        recoco_field_validation.add(ws[f"B{row}"])

    wb.save("sample.xlsx")


main()
